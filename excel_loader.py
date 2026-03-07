"""
excel_loader.py — Load leads from Excel/CSV without pandas.

This implementation uses `openpyxl` for .xlsx files and the standard
`csv` module for .csv files to avoid heavy `numpy`/`pandas` dependencies.
Expected columns: Name, Email, Company
"""

from typing import List, Dict, Any
from pathlib import Path
import csv
from openpyxl import load_workbook

from logger import get_logger

logger = get_logger("excel_loader")


def _normalize_header(h: str) -> str:
    return h.strip().title()


def load_leads_from_excel(filepath: str) -> List[Dict[str, Any]]:
    p = Path(filepath)
    if not p.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    leads: List[Dict[str, Any]] = []

    try:
        if filepath.lower().endswith(".csv"):
            with p.open(newline='', encoding='utf-8') as fh:
                reader = csv.DictReader(fh)
                cols = [_normalize_header(c) for c in reader.fieldnames or []]
                rows = list(reader)

                logger.info("Loaded CSV: %s (%d rows)", filepath, len(rows))

                for r in rows:
                    name = r.get('Name') or r.get('name') or r.get('NAME')
                    email = r.get('Email') or r.get('email') or r.get('EMAIL')
                    company = r.get('Company') or r.get('company') or r.get('COMPANY')
                    if not name or not email or not company:
                        continue
                    if '@' not in str(email):
                        continue
                    leads.append({
                        'name': str(name).strip(),
                        'email': str(email).strip().lower(),
                        'company': str(company).strip(),
                    })

        else:
            wb = load_workbook(filename=filepath, read_only=True, data_only=True)
            ws = wb.active

            # Read header
            rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header = next(rows)
            if not header:
                raise ValueError("Empty Excel file or missing header row")

            normalized = [_normalize_header(str(h)) if h is not None else "" for h in header]
            # Map column index to normalized name
            col_map = {i: name for i, name in enumerate(normalized)}

            required = {"Name", "Email", "Company"}
            if not required.issubset(set(normalized)):
                raise ValueError(f"Missing required columns. Found: {normalized}")

            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                count += 1
                values = {col_map[i]: (cell if cell is not None else "") for i, cell in enumerate(row)}
                name = str(values.get('Name', '')).strip()
                email = str(values.get('Email', '')).strip()
                company = str(values.get('Company', '')).strip()
                if not name or not email or not company:
                    continue
                if '@' not in email:
                    continue
                leads.append({'name': name, 'email': email.lower(), 'company': company})

            logger.info("Loaded XLSX: %s (%d rows scanned, %d valid)", filepath, count, len(leads))

    except Exception as exc:
        raise ValueError(f"Failed to read file {filepath}: {exc}")

    if not leads:
        raise ValueError("No valid leads found in the file.")

    logger.info("Loaded %d valid leads from: %s", len(leads), filepath)
    return leads
